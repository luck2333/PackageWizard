import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook
from collections import defaultdict


def compare_bga_files(result_file, reference_file, output_file):
    # 加载工作簿
    wb_result = load_workbook(result_file)
    wb_ref = load_workbook(reference_file)

    # 获取工作表
    sheet_result = wb_result.active
    sheet_ref = wb_ref.active

    # 创建高亮样式
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    bold_font = Font(bold=True)

    # 获取标题行
    headers = [cell.value for cell in sheet_result[1]]

    # 需要比较的列（原有列 + Exclude Pins）
    compare_columns = [
        'Pitch x (el)',
        'Pitch y (e)',
        'Number of pins along X',
        'Number of pins along Y',
        'Package Height (A)',
        'Standoff (A1)',
        'Body X (D)',
        'Body Y (E)',
        'Edge Fillet Radius',
        'Ball Diameter Normal (b)',
        'Exclude Pins'  # 新增列
    ]
    compare_columns = [col for col in compare_columns if col in headers]  # 过滤存在的列

    # 创建列索引映射
    col_indices = {header: idx + 1 for idx, header in enumerate(headers) if header in compare_columns}

    # 添加"错误数量"列（如果不存在）
    diff_count_col = len(headers) + 1
    if "错误数量" not in headers:
        sheet_result.cell(row=1, column=diff_count_col, value="错误数量").font = bold_font

    # 参考数据字典 {(PDF名称, 页码): 行数据}
    ref_data = {}
    for row in sheet_ref.iter_rows(min_row=2, values_only=True):
        key = (row[0], row[1])  # (PDF名称, 页码)
        ref_data[key] = row

    # 统计变量
    total_highlights = 0
    row_highlights = defaultdict(int)
    column_highlights = defaultdict(int)

    # 遍历结果文件并记录最大行号
    max_row = 1  # 至少有标题行
    for row_idx, row in enumerate(sheet_result.iter_rows(min_row=2, values_only=False), start=2):
        max_row = row_idx  # 更新最大行号
        pdf_name = row[0].value
        page_num = row[1].value
        ref_row = ref_data.get((pdf_name, page_num))
        if not ref_row:
            continue

        row_diff_count = 0

        # 比较每个目标列
        for col_name, col_idx in col_indices.items():
            result_val = row[col_idx - 1].value
            ref_val = ref_row[col_idx - 1]

            # 根据列类型选择比较方式
            if col_name == 'Exclude Pins':
                is_diff = not compare_exclude_pins(result_val, ref_val)
            else:
                is_diff = not compare_standard_columns(result_val, ref_val)

            if is_diff:
                row[col_idx - 1].fill = red_fill
                row_diff_count += 1
                column_highlights[col_name] += 1

        if row_diff_count > 0:
            sheet_result.cell(row=row_idx, column=diff_count_col, value=row_diff_count)
            total_highlights += row_diff_count
            row_highlights[row_idx] = row_diff_count

    # 新增：添加"错误数量"行（记录每列差异总数），固定在最后一行
    error_row = max_row + 1  # 最后一行的下一行
    # 在错误数量行添加标题
    sheet_result.cell(row=error_row, column=1, value="每列错误数量").font = bold_font
    # 填充每列的差异总数
    for col_name, col_idx in col_indices.items():
        sheet_result.cell(
            row=error_row,
            column=col_idx,
            value=column_highlights[col_name]
        ).font = bold_font  # 加粗显示列错误数

    # 保存结果
    wb_result.save(output_file)

    return {
        'total': total_highlights,
        'rows': dict(row_highlights),
        'columns': dict(column_highlights)
    }


def compare_exclude_pins(val1, val2):
    # 处理None或空值
    if val1 is None and val2 is None:
        return True
    if val1 is None or val2 is None:
        return False

    # 转换为字符串并标准化
    def standardize(value):
        value = str(value).strip()
        # 提取方括号内容并拆分为列表
        if value.startswith('[') and value.endswith(']'):
            content = value[1:-1].strip()
            # 分割逗号，过滤空值，统一小写
            items = [item.strip().lower() for item in content.split(',') if item.strip()]
            return sorted(items)  # 排序后比较
        return [value.lower()]

    # 比较标准化后的内容
    return standardize(val1) == standardize(val2)


def compare_standard_columns(val1, val2):
    """比较标准列（提取逗号分隔的中间值）"""

    def extract_middle(val):
        if val is None:
            return None
        if isinstance(val, str):
            parts = val.strip("[]").split(",")
            return parts[1].strip() if len(parts) >= 3 else None
        elif isinstance(val, list):
            return str(val[1]) if len(val) >= 3 else None
        return None

    mid1 = extract_middle(val1)
    mid2 = extract_middle(val2)

    # 处理None或数字比较
    if mid1 is None or mid2 is None:
        return mid1 == mid2
    try:
        return abs(float(mid1) - float(mid2)) < 1e-6
    except ValueError:
        return mid1.lower() == mid2.lower()


# 使用示例
if __name__ == "__main__":
    stats = compare_bga_files(
        result_file='bga_result_normalized.xlsx',
        reference_file='bga_standard_normalized.xlsx',
        output_file='bga_compare_result.xlsx'
    )
    print(f"\n结果已保存到 bga_compare_result.xlsx")