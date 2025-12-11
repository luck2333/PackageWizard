from openpyxl import load_workbook
from openpyxl.styles import Font


def normalize_data(text):
    """数据规范化函数"""
    if text is None:
        return None

    # 转换为字符串
    text = str(text)

    # 删除单引号、减号和.pdf
    text = text.replace("'", "").replace("-", "").replace(".pdf", "")

    # 处理Exclude Pins列的特殊格式
    if "Exclude Pins" in str(text) or ("[" in text and "]" in text):
        # 处理形式["[B1, C7]", [3.0, 3.0]] -> [B1,C7]
        if text.startswith('["[') and ']", [' in text:
            # 提取引脚部分（如"[B1, C7]"）
            start = text.find('["[') + 3
            end = text.find(']', start)
            pins = text[start:end]
            # 移除空格
            pins = pins.replace(" ", "")
            text = f"[{pins}]"

        # 处理形式["['A1', 'G18', 'H1']", '[3.0, 3.0, 3.0]', '-'] -> [A1,G18,H1]
        elif "[\"['" in text and "']\"" in text:
            # 提取引脚部分（如"['A1', 'G18', 'H1']"）
            start = text.find("[\"['") + 4
            end = text.find("']\"", start)
            pins = text[start:end]
            # 移除单引号和空格
            pins = pins.replace("'", "").replace(" ", "")
            text = f"[{pins}]"

        # 处理形式[None, []] -> [None]
        elif "[None, []" in text or "[ None, []" in text:
            text = "[None]"

        # 处理形式['None', '[]', '-'] -> [None]
        elif "['None', '[]', '-'" in text:
            text = "[None]"

        # 处理方括号格式中的"未检测出"
        elif "未检测出" in text:
            text = text.replace("未检测出", "None")
            # 将方括号中的空值替换为"None"
            text = text.replace("[, , ]", "[None, None, None]")

    return text


def normalize_excel_file(filename):
    """规范化Excel文件中的数据"""
    print(f"正在规范化文件: {filename}")

    # 加载工作簿
    wb = load_workbook(filename)
    ws = wb.active

    # 获取所有数据
    rows = list(ws.iter_rows())

    # 规范化所有单元格的数据
    for row in rows:
        for cell in row:
            if cell.value is not None:
                normalized_value = normalize_data(cell.value)
                cell.value = normalized_value

    # 保存规范化后的文件
    normalized_filename = filename.replace('.xlsx', '_normalized.xlsx')
    wb.save(normalized_filename)
    print(f"规范化完成，保存为: {normalized_filename}")

    return normalized_filename


# 规范化两个文件
normalized_file1 = normalize_excel_file('bga_result.xlsx')
normalized_file2 = normalize_excel_file('bga_standard.xlsx')
print("数据规范化完成！")