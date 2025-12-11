import os
import ast
import sys
import math
import time
import shutil  # 文件复制
import threading
import filetype  # 第三方库 filetype 1.2.0
import fitz
from PySide6.QtWidgets import QApplication, QMainWindow, \
    QHBoxLayout, QFileDialog, QMessageBox, QTableWidgetItem,QLabel

# 第三方库 PySide6 6.5.3
from PySide6.QtPdf import QPdfDocument
from PySide6.QtPdfWidgets import QPdfView
from PySide6.QtCore import QPoint, Qt, Signal, QEvent, QMargins, Slot
from PySide6.QtGui import QCloseEvent, QBrush, QColor, QCursor, QPen, QShortcut, QKeySequence
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
# 外部文件
from package_core.UI.Ui_MainWindowUI import Ui_MainWindow
from package_core.UI.ui_class import EnquirePopUp, Yes, No, MyGraphicsView, PuShButton_initial_qss, \
    Label_initial_qss, PuShButton_img_Draw_qss, PuShButton_list_Draw_qss, \
    Label_Draw_Img_qss, Label_Draw_List_qss, DetectThread, MyProgressDialog, RecoThread
from package_core.UI.Tools import create_dir, remove_dir, PDF_NAME, PDF_NAME_MINI

#外部后续添加文件
from package_core.Segment.Segment_function import get_type
# 常量
WIN_TITLE = "Package Wizard V1.1"
IMG_LABEL = "/{}个封装图"
LIST_LABEL = "/{}个封装表"
LABEL = "第{}/{}个封装图  第{}/{}个封装表"                 # pdf浏览器标签显示内容
LABEL_EMPTY = ""                    # 系统刚启动或刚加载pdf时显示
LABEL_DREW_IMG = "正在框封装图"          # 系统框选封装图时显示
LABEL_DREW_LIST = "正在框封装表"          # 系统框选封装表时显示
COMO_ITEMS = ['百分比']        # pdf浏览器下拉列表
TEMP_DIRECTORY = r"Result/temp"         # 临时存放pdf文件夹
TABLE_HEADERS = ['参数', 'Min', 'Type', 'Max']   # 尺寸数据表头
TABLE_GAP = 21             # 表格滑轮大小
TABLE_COLOR1 = QColor(207, 214, 236)
TABLE_COLOR2 = QColor(233, 236, 246)
# BGA_TABLE = ['实体长D', '实体宽E', '实体高A', '支撑高A1', '球直径b', '球行数', '球列数', '行球间距e', '列球间距e', '缺PIN否']
# QFN_TABLE = ['实体长D', '实体宽E', '实体高A', '支撑高A1', '端子高A3', 'PIN长L', 'PIN宽b', '行PIN数', '列PIN数', '行/列PIN间距e',
#              '散热盘长D2', '散热盘宽E2', '削角否', '削角长度', '端子圆角否']
# QFP_TABLE = ['实体长D1', '实体宽E1', '实体高A', '支撑高A1', '端子高A3', '外围长D', '外围宽E', 'PIN长L', 'PIN宽b', '行PIN数',
#              '列PIN数', '行/列PIN间距e', '散热盘长D2', '散热盘宽E2', '削角长度', '端子厚度', '接触角度', '端腿角度', '主体顶部绘制角度']
# SON_TABLE = ['实体长D', '实体宽E', '实体高A', 'PIN长L', 'PIN宽b', 'PIN行数', 'PIN列数', '行PIN数', '列PIN数',
#              'PIN_Pitche', 'PIN端距', 'layout建议值']
# SOP_TABLE = ['实体长D', '实体宽E', '实体高A', 'PIN长L', 'PIN宽b', 'PIN行数', 'PIN列数', '行PIN数', '列PIN数',
#              'PIN_Pitche', 'PIN端距', 'layout建议值']
BGA_TABLE = ['Pitch x (el)', 'Pitch y (e)', 'Number of pins along X', 'Number of pins along Y',
             'Package Height (A)', 'Standoff (A1)', 'Body X (D)', 'Body Y (E)', 'Edge Fillet Radius',
             'Ball Diameter Normal (b)', 'Exclude Pins']
QFN_TABLE = ['Pitch x (el)', 'Pitch y (e)', 'Number of pins along X', 'Number of pins along Y',
             'Package Height (A)', 'Standoff (A1)', 'Pull Back (p)', 'Body X (D)', 'Body Y (E)',
             'Lead style', 'Pin Length (L)', 'Lead width (b)', 'Lead Height (c)', 'Exclude Pins',
             'Thermal X (D2)', 'Thermal Y (E2)']
QFP_TABLE = ['Number of pins along X', 'Number of pins along Y', 'Package Height (A)', 'Standoff (A1)',
             'Span X (D)', 'Span Y (E)', 'Body X (D1)', 'Body Y (E1)', 'Body draft (θ)', 'Edge Fillet radius',
             'Lead Length (L)', 'Lead width (b)', 'Lead Thickness (c)', 'Lead Radius (r)', 'Thermal X (D2)', 'Thermal Y (E2)']
SON_TABLE = ['Pitch (e)', 'Number of pins', 'Package Height (A)', 'Standoff (A1)', 'Pull Back (p)', 'Body X (D)',
             'Body Y (E)', 'Lead style', 'Lead Length (L)', 'Lead width (b)', 'Lead Height (c)', 'Exclude Pins', 'Thermal X (D2)', 'Thermal Y (E2)']
SOP_TABLE = ['Pitch(e)', 'number of pins', 'Package Height (A)', 'Standoff (A1)', 'Span X (E)',
             'Body X (E1)', 'Body Y (D)', 'Body draft (θ)', 'Edge fillet radius',
             'Lead Length (L)', 'Lead width (b)', 'Lead Thickness (c)', 'Lead Radius (r)', 'Thermal X (E2)',
             'Thermal Y (D2)']

PACKAGE_COLOR = (1, 0, 0)
KEYVIEW_COLOR = (0, 1, 1)
TOP_COLOR = (0.5, 1, 0)
SIDE_COLOR = (1, 0.5, 0)
DETAIL_COLOR = (0.5, 0, 1)
FORM_COLOR = (0, 0, 1)
NOTE_COLOR = (1, 0, 0.5)

ZOOM_MULTIPLIER = 1.2          # pdf浏览器界面每次缩放比例

class MyWindow(QMainWindow):
    """
    主界面
    """
    def __init__(self, parent=None):
        super(MyWindow, self).__init__(parent)
        # 建立ui对象
        self.ui = Ui_MainWindow()
        # 显示ui
        self.ui.setupUi(self)

        # 设置下拉框内容
        self.ui.comboBox.addItems(COMO_ITEMS)

        self.color = (1, 0, 0)  # 默认框图颜色 红色
        # pdf模式伸缩比
        self.pdf_factor = 1
        # 基底缩放比, 为缩放基础
        self.pdf_factor_is_one = self.pdf_factor
        # pdf显示伸缩比
        self.pdf_view_factor = 1

        # 当前模式pdf页面展示宽
        self.pdf_show_width = 0
        # 当前模式pdf页面展示高
        self.pdf_show_height = 0

        # 当前模式颠倒页pdf显示宽
        self.trans_pdf_show_width = 0
        # 当前模式颠倒页pdf显示高
        self.trans_pdf_show_height = 0
        # pdf页面信息{页码:[是否颠倒，长，宽，长宽比，前面几个颠倒页]},现在先默认一个文档里pdf的页面宽高相同，及现在记录的数据只有时候颠倒和颠倒页数记录有用
        self.pdf_message = {}
        # 存储pdf页数
        self.pdf_page_count = 0
        # 存储第一个颠倒页的页码
        self.first_trans_page_number = -1
        # 存储各个分辨率下的右侧空白,将区分页面颠倒
        self.pdf_factor_right_width = {True: {}, False: {}}
        # 是否具有颠倒页
        self.pdf_have_trans = False

        # pdf文档首页的宽高
        self.pdf_height = 0
        self.pdf_width = 0
        self.width_height = 0

        # qt和屏幕的像素比
        self.device_pixel = 0
        # 屏幕分辨率
        self.screen_width = 0
        # 保存屏幕宽度基底
        self.base_screen_width = self.screen_width
        # 屏幕基准分辨率
        self.screen_base = 1920
        self.ui_width = 0
        self.ui_height = 0
        # self.label_text = ""

        # 设置文档上下左右边距
        self.pdf_margin = 10
        self.ui.pdfView.setDocumentMargins(QMargins(self.pdf_margin, self.pdf_margin, self.pdf_margin, self.pdf_margin))

        # 根据屏幕比例计算伸缩比
        self.cal_factor()

        # 当前pdf的最大最小缩放比例
        self.max_factor = 0
        self.min_factor = 0
        # 动态缩放比例计算因子
        self.max_factor_cal = pow(ZOOM_MULTIPLIER, 2)
        self.min_factor_cal = pow(ZOOM_MULTIPLIER, 2)
        # 存储计算放大缩小时进行的比例操作数据
        self.in_or_out = 0

        # 单页垂直方向距离
        self.page_vertical = self.pdf_margin
        # 水平方向举例
        self.page_horizontal = 0

        # pdf浏览器宽高
        self.pdf_view_width = 0
        self.pdf_view_height = 0
        # 当前pdf展示宽高
        self.document_screen_width = 0

        # 表格宽
        self.table_width = 0

        """pdf路径"""
        self.temp_pdf = ""  # 使用pdf
        self.pdf_path = ""  # 后续操作使用pdf
        self.ace_pdf = ""   # 最开始pdf
        self.history_path = "c:\\"   # 历史路径

        # 当前操作的页码
        self.show_page_number = 1

        # 设置快捷键
        self.esc = QShortcut(QKeySequence("Escape"), self)

        """pdf浏览器设置"""
        self.m_document = QPdfDocument(self)  # pdf文件
        self.ui.pdfView.setDocument(self.m_document)
        self.nav = self.ui.pdfView.pageNavigator()    # 设置pdf页面导航
        self.ui.pdfView.setZoomMode(QPdfView.ZoomMode.Custom)   # 默认pdf浏览大小模式
        self.ui.pdfView.setPageMode(QPdfView.PageMode.MultiPage)   # 设置多页模式
        self.ui.comboBox.setCurrentIndex(0)         # 设置pdf浏览大小模式 百分比

        """缩略图设置"""
        self.m_document_mini = QPdfDocument(self)
        self.ui.pdfView_thumbnail.setDocument(self.m_document_mini)
        self.nav_mini = self.ui.pdfView_thumbnail.pageNavigator()
        self.ui.pdfView_thumbnail.setZoomMode(QPdfView.ZoomMode.Custom)
        self.ui.pdfView_thumbnail.setPageMode(QPdfView.PageMode.MultiPage)

        """pdf浏览器绘图层"""
        self.horizontalLayout_11 = QHBoxLayout(self.ui.pdfView)    # 添加布局
        self.graphicsView = MyGraphicsView(self.ui.pdfView)        # 添加画板
        self.horizontalLayout_11.setSpacing(0)
        self.horizontalLayout_11.setContentsMargins(0, 0, 0, 0)      # 布局格式修改
        self.graphicsView.setStyleSheet(u"background-color: rgba(255, 255, 255, 0)")  # 修改画板样式
        self.horizontalLayout_11.addWidget(self.graphicsView)   # 布局添加画板

        """控件动作连接响应函数"""
        self.nav.currentPageChanged.connect(self.show_page)       # pdf浏览器页数改变
        self.graphicsView.mouse_released[list].connect(self.process_rect_data) # 框选结束 转换坐标
        self.ui.pushButton_open.clicked.connect(self.open_file)     # 打开文件
        self.ui.pushButton_detect.clicked.connect(self.app_detect)   # 自动搜索封装图对象
        self.ui.pushButton_reco.clicked.connect(self.app_reco)      # 封装图或封装表提取参数
        self.ui.pushButton_edit.clicked.connect(self.app_edit)      # 尺寸信息表格编辑
        self.ui.pushButton_save.clicked.connect(self.app_save)      # 尺寸信息保存
        self.ui.pushButton_pre.clicked.connect(self.package_pre)    # 上一个封装对象
        self.ui.pushButton_next.clicked.connect(self.package_next)   # 下一个封装对象
        self.ui.pushButton_zoomout.clicked.connect(self.pdf_zoom_out)   # pdf浏览器放大
        self.ui.pushButton_zoomin.clicked.connect(self.pdf_zoom_in)    # pdf浏览器缩小
        self.ui.lineEdit_page.returnPressed.connect(self.edit_page_jump)  # pdf浏览器页面跳转
        self.ui.lineEdit_choose_img.returnPressed.connect(self.img_jump)  # 封装图页面跳转
        self.ui.lineEdit_choose_list.returnPressed.connect(self.list_jump) # 封装表页面跳转
        self.ui.comboBox.currentIndexChanged.connect(self.pdf_combo_show) # pdf浏览器大小模式切换
        self.ui.pushButton_draw_img.clicked.connect(self.get_img_rect_data) # 框封装图
        self.ui.pushButton_draw_list.clicked.connect(self.get_list_rect_data)  # 框封装表
        self.esc.activated.connect(self.rect_setting_restore)    # esc取消框选
        self.screen().physicalSizeChanged.connect(self.screen_change)   # 屏幕分辨率改变
        self.screen().geometryChanged.connect(self.screen_change)    # 屏幕分辨率改变
        self.ui.pdfView.verticalScrollBar().valueChanged.connect(self.page_change_with_bar) # 屏幕滚动

        create_dir(TEMP_DIRECTORY)      # 若temp文件夹存在，则清空，否则创建

        self.setup()   # 界面参数初始化

    def cal_factor(self):
        """
        获取屏幕分辨率并计算伸缩比
        :return:
        """
        self.pdf_factor = 1
        # 获取当前屏幕分辨率
        self.get_screen_width()
        # 保存屏幕宽度基底
        self.base_screen_width = self.screen_width
        # 屏幕分辨率过高或过低，降低默认伸缩比
        if self.screen_width > self.screen_base:
            self.pdf_factor = math.floor(self.screen_base / self.screen_width * 10 - 1) / 10
        elif self.screen_width < self.screen_base:
            self.pdf_factor = math.floor(self.screen_width / self.screen_base * 10 - 1) / 10
        # 设置比例1标志, 为缩放基础
        self.pdf_factor_is_one = self.pdf_factor

    def screen_change(self):
        """
        屏幕分辨率改变
        :return:
        """
        # 重新根据屏幕分辨率计算伸缩比
        self.cal_factor()

        # 重新获取pdf浏览器的宽高
        self.pdf_view_width = self.ui.pdfView.size().width()  # pdf浏览器实时宽度
        self.pdf_view_height = self.ui.pdfView.size().height()  # pdf浏览器实时高度

        # 重新打开pdf
        self.load_pdf(self.temp_pdf, self.show_page_number)

    def page_change_with_bar(self):
        """
        进行页面跳转
        :return:
        """
        # 当前显示页码
        self.show_page_number = self.nav.currentPage()
        # 设置跳转
        self.ui.lineEdit_page.setText(f"{self.show_page_number + 1}")

    def load_pdf(self, pdf, show_number):
        """
        重新打开pdf
        :param pdf:打开pdf路径
        :param show_number:显示的pdf页面
        :return:
        """
        # 关闭画布
        if self.graphicsView.isVisible():
            # 重置参数
            self.draw_img = 0
            self.draw_list = 0
            # 关闭画布
            self.draw_add_func_restore()
        # 清空记录
        self.pdf_factor_right_width[True].clear()
        self.pdf_factor_right_width[False].clear()
        if self.pdf_width > 596:
            # 计算当前pdf的打开分辨率
            self.pdf_factor = math.floor(596 / self.pdf_width * 10 - 1) / 10 * self.pdf_factor_is_one
        else:
            # 如果有颠倒页为了适配，将默认比例缩小
            if self.pdf_have_trans:
                self.pdf_factor = self.pdf_factor_is_one * self.width_height
            else:
                # 使用当前分辨率基底分辨率
                self.pdf_factor = self.pdf_factor_is_one
        # 动态设置当前pdf的最大最小缩放比例
        self.max_factor = self.pdf_factor * self.max_factor_cal
        self.min_factor = self.pdf_factor / self.min_factor_cal

        # 设置默认模式选择
        self.ui.comboBox.setCurrentIndex(0)
        # 重新加载pdf
        self.m_document.load(pdf)

        # 设置默认缩放比例
        self.ui.pdfView.setZoomFactor(self.pdf_factor)
        # 设置页码显示
        self.ui.lineEdit_page.setText(f"{show_number + 1}")
        self.ui.label_total_page.setText(f"/{self.m_document.pageCount()}")
        # 跳转第二页，计算pdf显示高，只有页码大于1进行计算
        if self.pdf_page_count > 1:
            self.nav.jump(1, QPoint())
            # 得到pdf的展示height
            get_height = threading.Thread(target=self.get_pdf_show_height, args=(show_number, True,))
            get_height.start()

    def setup(self):
        """
          界面参数重置
        :return:
        """
        self.ui.label_img_list.setVisible(0)      # 隐藏底部框图状态栏
        # self.ui.pdfView_thumbnail.setStyleSheet("""background-color:red;""")
        self.ui.tableWidget.setRowCount(0)  # 表格内容清空
        self.ui.lineEdit_type.setText('')    # 封装类型清空

        # self.ui.label_package.setText(LABEL.format('-', '0', '-', '0'))
        self.ui.lineEdit_choose_img.setText('—')
        self.ui.lineEdit_choose_list.setText('—')
        self.ui.label_total_img.setText(IMG_LABEL.format(0))
        self.ui.label_total_list.setText(LIST_LABEL.format(0))

        """界面部分功能锁定"""
        self.ui.pushButton_pre.setEnabled(0)
        self.ui.pushButton_next.setEnabled(0)
        self.ui.pushButton_detect.setEnabled(0)
        self.ui.pushButton_reco.setEnabled(0)
        self.ui.pushButton_draw_img.setEnabled(0)
        self.ui.pushButton_draw_list.setEnabled(0)
        self.ui.pushButton_edit.setEnabled(0)
        self.ui.pushButton_save.setEnabled(0)
        self.ui.lineEdit_type.setEnabled(0)
        self.ui.tableWidget.setEnabled(0)
        self.ui.pushButton_draw_img.setVisible(0)
        self.ui.pushButton_draw_list.setVisible(0)
        # self.ui.pdfView_thumbnail.setVisible(0)
        # self.ui.frame_nav_get.setVisible(0)
        self.ui.pdfView_thumbnail.setDocument(None)
        self.graphicsView.setVisible(False)  # 画板隐藏

        """框图相关设置"""
        self.draw_img = 0
        self.draw_list = 0

        """其他信息"""
        self.package = [] # 可以分为两大类(封装图，封装表)，3小类(检测的封装图，框选的封装图，框选的封装表)
            # 若是封装图'img' ->{'page', 'type', 'rect', 'new_rect', 'package_type', 'part_content': [{}, {}], 'reco_content'}
            #         -> {'page', 'type', 'rect', 'new_rect', 'package_type', 'part_content': [], 'reco_content'}
            # 若是封装表'list' ->{'page', 'type', 'rect', 'new_rect', 'package_type', 'reco_content'}
        self.type_dict = {}
        self.current = 0        # 当前展示封装信息 索引
        self.show_page_number = 0  # 显示页重置


    def get_screen_width(self):
        """
        获取当前屏幕的宽
        :return:
        """
        # qt和屏幕的像素比
        self.device_pixel = self.screen().devicePixelRatio()
        # 获取当前屏幕范围
        screen_rect = self.screen().geometry().getRect()
        # 判断获取屏幕的宽
        self.screen_width = screen_rect[2] * self.device_pixel

    def get_pdf_show_height(self, page_show, is_jump):
        """
        得到pdf当前展示高度
        :param page_show: 跳转操作，页码，或者右侧滚动条
        :param is_jump: True 跳页码，False右侧滚动条
        :return:
        """
        # 沉睡
        time.sleep(0.001)
        # 计算当前pdf展示高
        self.pdf_show_height = (self.ui.pdfView.verticalScrollBar().value() - self.pdf_margin -
                                self.ui.pdfView.pageSpacing())
        # 获取显示宽
        self.pdf_show_width = self.pdf_show_height * self.width_height
        if self.pdf_show_width != 0:
            # 获取当前的缩放比
            self.pdf_view_factor = (self.pdf_width / self.pdf_show_width)
        # 如果有颠倒页
        if self.pdf_have_trans:
            self.nav.jump(self.first_trans_page_number + 1, QPoint())
            # 得到pdf的颠倒页height
            get_height = threading.Thread(target=self.get_pdf_trans_show_height, args=(page_show, is_jump,))
            get_height.start()
        else:
            if is_jump:
                # 跳到指定页
                self.nav.jump(page_show, QPoint())
            else:
                self.ui.pdfView.verticalScrollBar().setValue(page_show)

    def get_pdf_trans_show_height(self, page_show, is_jump):
        """
        得到pdf当前展示高度
        :param page_show: 跳转操作，页码，或者右侧滚动条
        :param is_jump: True 跳页码，False右侧滚动条
        :return:
        """
        # 沉睡
        time.sleep(0.001)
        # 计算当前pdf展示高
        self.trans_pdf_show_width = (self.ui.pdfView.verticalScrollBar().value() - self.pdf_margin -
                                     self.ui.pdfView.pageSpacing() * (
                                             self.first_trans_page_number + 1) - self.first_trans_page_number * self.pdf_show_height)
        # 获取显示宽
        self.trans_pdf_show_height = self.trans_pdf_show_width * self.pdf_message[self.first_trans_page_number][3]
        if self.pdf_show_width != 0:
            # 获取当前的缩放比
            self.pdf_view_factor = (self.pdf_width / self.pdf_show_width)
        if is_jump:
            # 跳到指定页
            self.nav.jump(page_show, QPoint())
        else:
            self.ui.pdfView.verticalScrollBar().setValue(page_show)

    def open_file(self):
        """
            打开新文档并重置系统参数
        :return:
        """
        name, _ = QFileDialog.getOpenFileNames(self, "选择pdf文件", self.history_path, '*.pdf')
        if len(name) != 0:
            self.pdf_path = name[0]    # 获取pdf路径
            kind = filetype.guess(self.pdf_path)
            if kind and (kind.extension == 'pdf'):
                try:
                    self.pdf_message.clear()       # 清空页面方向记录
                    self.pdf_have_trans = False   # 清空颠倒页标志
                    self.first_trans_page_number = -1   # 清空第一个颠倒页的页码
                    trans_num = 0             # 记录当前页面之前有几个横向页面
                    with fitz.open(self.pdf_path) as doc:
                        # 记录页数
                        self.pdf_page_count = doc.page_count
                        # 遍历记录数据{页码:[是否颠倒，宽，高，宽高比，前面几个颠倒页]}
                        for page_number in range(0, doc.page_count):
                            # 判断颠倒
                            if doc[page_number].rect[2] >= doc[page_number].rect[3]:
                                # 说明有颠倒页
                                self.pdf_have_trans = True
                                # 记录颠倒页页码
                                if self.first_trans_page_number == -1:
                                    self.first_trans_page_number = page_number
                                # 记录信息
                                self.pdf_message[page_number] = [True, doc[page_number].rect[2],
                                                                 doc[page_number].rect[3],
                                                                 doc[page_number].rect[2] / doc[page_number].rect[3],
                                                                 trans_num]
                                # 记录颠倒页
                                trans_num += 1
                            else:
                                # 记录信息
                                self.pdf_message[page_number] = [False, doc[page_number].rect[2],
                                                                 doc[page_number].rect[3],
                                                                 doc[page_number].rect[2] / doc[page_number].rect[3],
                                                                 trans_num]
                        # 获取pdf首页
                        self.pdf_width = doc[0].rect[2]  # 页宽
                        self.pdf_height = doc[0].rect[3]  # 页高
                        # 无论怎么缩放pdf的长宽比不变
                        self.width_height = self.pdf_width / self.pdf_height

                    self.setup()       # 界面参数重置
                    path_split = self.pdf_path.split('/')      # 切割路径 不规范写法
                    pdf_name = path_split[-1]           # 获取pdf_name
                    self.history_path = '/'.join(path_split[0:-1])      # 存储历史路径
                    self.setWindowTitle(WIN_TITLE + '———' + str(pdf_name))   # 设置窗口标题

                    self.temp_pdf = TEMP_DIRECTORY + "\\" + pdf_name
                    shutil.copy2(self.pdf_path, self.temp_pdf)   # 创建副本
                    self.ace_pdf = self.pdf_path  # 最开始pdf地址 禁止对该路径的pdf做任何修改
                    self.pdf_path = self.temp_pdf   # 给后端传递pdf地址

                    self.load_pdf(self.temp_pdf, 0)     # 加载pdf

                    self.ui.pushButton_detect.setEnabled(1)  # 释放自动搜索按钮
                    self.ui.pushButton_draw_img.setEnabled(1)   # 释放框图按钮
                    self.ui.pushButton_draw_list.setEnabled(1)   # 释放框表按钮
                except Exception as e:
                    QMessageBox.critical(self, 'Failed to open', e.__str__())
            else:
                QMessageBox.critical(self, 'Not a pdf', '这似乎不是一篇pdf')

    def app_detect(self):
        """
            开启多线程，前端弹窗，后端进行yolox自动搜索
        :return:
        """
        # 禁用相关按钮
        self.ui.pushButton_open.setEnabled(0)
        self.ui.pushButton_detect.setEnabled(0)
        self.ui.pushButton_reco.setEnabled(0)
        self.ui.pushButton_pre.setEnabled(0)
        self.ui.pushButton_next.setEnabled(0)
        self.ui.pushButton_draw_img.setEnabled(0)
        self.ui.pushButton_draw_list.setEnabled(0)
        self.ui.pushButton_edit.setEnabled(0)
        self.ui.pushButton_save.setEnabled(0)

        self.progress_dialog = MyProgressDialog(self, '页面前处理筛选', '页面筛选中')
        self.progress_dialog.pushButton.clicked.connect(self.kill_thread)
        self.thread = DetectThread(self, self.pdf_path)
        self.thread.signal_end_page.connect(self.end_fir_process)

        self.thread.signal_end.connect(self.process_detect_data)

        self.thread.start()
        # 在 MyWindow.app_detect 中连接信号
        self.thread.signal_error.connect(self.on_detect_error)

    # 创建新的槽函数来处理错误
    @Slot(str)
    def on_detect_error(self, error_message):
        self.close_progress_dialog()  # 同样需要关闭进度条
        QMessageBox.critical(self, '自动搜索流程出现错误', error_message)
        # 恢复UI按钮状态等
        # 解除禁用相关按钮
        self.ui.pushButton_open.setEnabled(1)
        self.ui.pushButton_detect.setEnabled(1)
        # 连接信号
        self.thread.signal_end.connect(self.process_detect_data)

    # # 修改槽函数签名以接收数据
    # @Slot(list)  # 明确指定接收的参数类型
    # def process_detect_data(self, package_data):  # 直接接收数据
    #     self.close_progress_dialog()
    #     if (len(package_data)):
    #         # 直接使用接收到的 package_data
    #         self.package = package_data

    def app_reco(self):
        """
        参数识别
        :return:
        """
        # 禁用相关按钮
        self.ui.pushButton_open.setEnabled(0)
        self.ui.pushButton_detect.setEnabled(0)
        self.ui.pushButton_reco.setEnabled(0)
        self.ui.pushButton_pre.setEnabled(0)
        self.ui.pushButton_next.setEnabled(0)
        self.ui.pushButton_draw_img.setEnabled(0)
        self.ui.pushButton_draw_list.setEnabled(0)
        self.ui.pushButton_edit.setEnabled(0)
        self.ui.pushButton_save.setEnabled(0)

        self.progress_dialog = MyProgressDialog(self, '封装图信息识别', '参数信息识别中')
        self.progress_dialog.pushButton.clicked.connect(self.kill_thread)

        # 先获取当前要识别的封装包数据
        current_package_data = self.package[self.current - 1]
        # 从封装包数据中获取它所在的正确页码
        package_page_num = current_package_data['page']
        # 使用正确的页码（package_page_num）来初始化 RecoThread
        self.detect = RecoThread(self, self.pdf_path, package_page_num, current_package_data, self.package, self.type_dict)
        # self.detect = RecoThread(self, self.pdf_path,self.nav.currentPage(),self.package[self.current-1],self.package,self.type_dict) # 传入self.package ,self.nav.currentPage(),+1为实际PDF页

        self.detect.signal_end.connect(self.process_reco_data)

        self.detect.start()

    def app_edit(self):
        """
        点击编辑控件，右边控件显示可编辑状态
        """
        self.ui.lineEdit_type.setEnabled(1)
        self.ui.lineEdit_type.setFocus()
        self.ui.tableWidget.setEnabled(1)

    def app_save(self):
        """
        对表格中数据进行保存
        """
        QMessageBox.information(self, '保存功能', '功能持续开发中')

    def end_fir_process(self):
        """前处理筛选结束开始进行自动搜索"""
        self.close_progress_dialog()
        self.thread.resume()

        # self.progress_dialog = MyProgressDialog(self, '自动搜索封装图', '自动搜索中', 1) # 绑定专有定时器
        # self.progress_dialog.progressBar.setTextVisible(1)
        # self.progress_dialog.pushButton.clicked.connect(self.kill_thread)
        # # 给自动搜索进度条绑定专用定时器
        # self.progress_dialog.timer.timeout.connect(self.update_detect_process)
        # self.progress_dialog.timer.start(300)

        self.progress_dialog = MyProgressDialog(self, '自动搜索封装图', '自动搜索中...')
        self.progress_dialog.pushButton.clicked.connect(self.kill_thread)

    def kill_thread(self):
        """杀死线程"""
        self.thread.terminate()
        self.close_progress_dialog()

        # 释放按钮
        self.ui.pushButton_open.setEnabled(1)
        self.ui.pushButton_detect.setEnabled(1)
        self.ui.pushButton_reco.setEnabled(1)
        self.ui.pushButton_pre.setEnabled(1)
        self.ui.pushButton_next.setEnabled(1)
        self.ui.pushButton_draw_img.setEnabled(1)
        self.ui.pushButton_draw_list.setEnabled(1)

    def close_progress_dialog(self):
        """关闭动态进度条弹窗"""
        self.progress_dialog.timer.stop()
        self.progress_dialog.timer_label.stop()
        if (self.progress_dialog.isVisible()):
            self.progress_dialog.close()


    """识别函数的附加功能 -> 尺寸信息展示：表格选择+内容填充"""
    @Slot(int)
    def process_reco_data(self):
        """处理识别结果"""
        # 处理数据
        self.package[self.current - 1]['reco_content'] = self.detect.result
        # 填充表格
        self.show_content()
        self.close_progress_dialog()
        # 处理识别结果
        print(f"识别结束")

        # 恢复相关按钮
        self.ui.pushButton_open.setEnabled(1)
        self.ui.pushButton_detect.setEnabled(1)
        self.ui.pushButton_reco.setEnabled(1)
        self.ui.pushButton_pre.setEnabled(1)
        self.ui.pushButton_next.setEnabled(1)
        self.ui.pushButton_draw_img.setEnabled(1)
        self.ui.pushButton_draw_list.setEnabled(1)

    def show_content(self):
        """
            表格内容填充
        :return:
        """
        self.ui.pushButton_edit.setEnabled(1)
        self.ui.pushButton_save.setEnabled(1)
        # 有数据有封装类型
        # 有数据无封装类型
        # 无数据有封装类型
        # print(self.package[self.current - 1])
        data_content = self.package[self.current - 1]['reco_content']
        # data = [sublist[1:] if i != len(data_content) - 1 else sublist for i, sublist in enumerate(data_content)]
        # #------------------------------
        # # Excel文件路径
        # file_path = "output.xlsx"
        #
        # # 检查文件是否存在
        # if os.path.exists(file_path):
        #     # 加载现有工作簿
        #     wb = load_workbook(file_path)
        #     ws = wb.active
        #     # 找到最后一行（数据从第2行开始）
        #     last_row = ws.max_row + 1
        # else:
        #     # 创建新工作簿
        #     wb = Workbook()
        #     ws = wb.active
        #     # 写入表头
        #     for col_num, header in enumerate(BGA_TABLE, start=2):
        #         ws.cell(row=1, column=col_num, value=header)
        #     last_row = 2  # 从第2行开始写数据
        #
        # # 写入新数据（追加到最后一行）
        # for col_num, sublist in enumerate(data, start=2):
        #     # 合并子列表为字符串（用逗号分隔）
        #     cell_value = ', '.join(str(item) for item in sublist)
        #     ws.cell(row=last_row, column=col_num, value=cell_value)
        #
        # # 填充PDF名称（自动编号）
        # pdf_name = f"{os.path.basename(self.pdf_path.split('/')[1])}"
        # ws.cell(row=last_row, column=1, value=pdf_name)
        #
        # # 保存Excel文件
        # wb.save(file_path)
        #------------------------------
        row_count = len(data_content)  # 行数
        if (self.ui.tableWidget.rowCount() > 1):  # 有数据 有封装类型
            # column_count = len(data_content[0])  # 列数
            for i in range(row_count):
                for j in range(len(data_content[i])):
                    if (j != 0):
                        # self.ui.tableWidget.item(i, j).setText(str(data_content[i][j]))  # 内容填写
                        item = self.ui.tableWidget.item(i, j)
                        # 处理最后一行的第2列（j == 1）
                        if i == row_count - 1 and j == 1:
                            try:
                                pin_list = ast.literal_eval(data_content[i][1])
                                value_list = ast.literal_eval(data_content[i][2])
                                if not isinstance(pin_list, list) or not isinstance(value_list, list):
                                    raise ValueError("解析失败：不是列表")
                                else:
                                    self.ui.tableWidget.setSpan(row_count - 1, 1, 1, 3)
                            except Exception as e:
                                print("解析失败：", e)
                                # 如果失败就正常填充
                                item = QTableWidgetItem(str(data_content[i][j]))
                                item.setTextAlignment(Qt.AlignCenter)
                                self.ui.tableWidget.setItem(i, j, item)
                                continue

                            # 构造彩色富文本
                            color_texts = []
                            for pin, val in zip(pin_list, value_list):
                                if val == 3.0:
                                    color = "purple"
                                elif val == 1.0:
                                    color = "green"
                                elif val == 0.0:
                                    color = "red"
                                else:
                                    color = "black"
                                color_texts.append(f'<font color="{color}">{pin}</font>')

                            html_text = ", ".join(color_texts)

                            label = QLabel()
                            label.setTextFormat(Qt.RichText)
                            label.setText(html_text)
                            label.setAlignment(Qt.AlignCenter)
                            label.setWordWrap(True)

                            self.ui.tableWidget.setCellWidget(i, j, label)
                        else:
                            item = QTableWidgetItem(str(data_content[i][j]))
                            item.setTextAlignment(Qt.AlignCenter)
                            self.ui.tableWidget.setItem(i, j, item)
            # self.ui.tableWidget.resizeRowsToContents()
        else:
            # 生成表格同时填充数据
            self.ui.tableWidget.setRowCount(row_count)  # 设置单元格行数
            for i in range(row_count):
                if (i % 2):
                    color = TABLE_COLOR1
                else:
                    color = TABLE_COLOR2
                for j in range(len(TABLE_HEADERS)):
                    cell = QTableWidgetItem()
                    cell.setBackground(QBrush(color))
                    cell.setTextAlignment(Qt.AlignCenter)
                    # cell.setText(str(data_content[i][j]))
                    self.ui.tableWidget.setItem(i, j, cell)
            for i in range(row_count):
                for j in range(len(data_content[i])):
                    self.ui.tableWidget.item(i, j).setText(str(data_content[i][j]))  # 内容填写
                    # item = self.ui.tableWidget.item(i, j)
                    # item.setText(str(data_content[i][j]))  # 内容填写
                    # if i == row_count - 1:
                    #     item.setForeground(QBrush(QColor("red")))  # 设置字体颜色为红色


    def generate_table(self, package_type):
        """根据不同的封装类型生成不同表格"""
        # 保险起见，先清空表格内容
        self.ui.tableWidget.setRowCount(0)
        if (package_type == 'BGA'):
            row_count = len(BGA_TABLE)
            parameter = BGA_TABLE
        elif (package_type == 'QFN'):
            row_count = len(QFN_TABLE)
            parameter = QFN_TABLE
        elif (package_type == 'QFP'):
            row_count = len(QFP_TABLE)
            parameter = QFP_TABLE
        elif (package_type == 'DFN_SON' or package_type == 'DFN' or package_type == 'SON'):
            row_count = len(SON_TABLE)
            parameter = SON_TABLE
        elif (package_type == 'SOP'):
            row_count = len(SOP_TABLE)
            parameter = SOP_TABLE
        else:
            row_count = 0
        self.ui.tableWidget.setRowCount(row_count)  # 设置单元格行数
        for i in range(row_count):
            if (i % 2):
                color = TABLE_COLOR1
            else:
                color = TABLE_COLOR2
            for j in range(len(TABLE_HEADERS)):
                cell = QTableWidgetItem()
                cell.setBackground(QBrush(color))
                cell.setTextAlignment(Qt.AlignCenter)
                if (j == 0):
                    cell.setText(parameter[i])
                self.ui.tableWidget.setItem(i, j, cell)
        # if (package_type == 'BGA'):
        #     row_count = 10
        #     parameter = BGA_TABLE
        # elif (package_type == 'QFN'):
        #     row_count = 15
        #     parameter = QFN_TABLE
        # elif (package_type == 'QFP'):
        #     row_count = 19
        #     parameter = QFP_TABLE
        # elif (package_type == 'DFN_SON' or package_type == 'DFN' or package_type == 'SON'):
        #     row_count = 12
        #     parameter = SON_TABLE
        # elif (package_type == 'SOP'):
        #     row_count = 12
        #     parameter = SOP_TABLE
        # else:
        #     row_count = 0
        # self.ui.tableWidget.setRowCount(row_count)    # 设置单元格行数
        # for i in range(row_count):
        #     if (i % 2):
        #         color = TABLE_COLOR1
        #     else:
        #         color = TABLE_COLOR2
        #     for j in range(len(TABLE_HEADERS)):
        #         cell = QTableWidgetItem()
        #         cell.setBackground(QBrush(color))
        #         cell.setTextAlignment(Qt.AlignCenter)
        #         if (j == 0):
        #             cell.setText(parameter[i])
        #         self.ui.tableWidget.setItem(i, j, cell)
        # 合并单元格
        # if (package_type == 'BGA'):
        #     self.ui.tableWidget.setSpan(row_count - 1, 1, 1, 3)
        # elif (package_type == 'QFN'):
        #     self.ui.tableWidget.setSpan(row_count - 3, 1, 1, 3)
        #     self.ui.tableWidget.setSpan(row_count - 1, 1, 1, 3)

    """检测函数的附加功能 -> 绑定自动搜索进度条 + 自动搜索数据处理 + 封装图切换 + pdf画框"""
    def update_detect_process(self):
        """根据当前处理页获取检测进度条值"""
        if (self.progress_dialog.progress_value < 100):
            self.progress_dialog.progress_value = \
                int(self.thread.PreProcess.current_page / len(self.thread.page_list) * 100) if self.thread.PreProcess is not None else 0
        self.progress_dialog.progressBar.setValue(self.progress_dialog.progress_value)

    @Slot(list)
    def process_detect_data(self, package_list):
        """
        自动搜索流程结束，处理并展示检测结果。
        该函数作为总协调，调用辅助函数来处理具体任务。
        """
        self.close_progress_dialog()
        # 1. 更新内部数据模型
        self.package = package_list if package_list is not None else []
        # 2. 【可选逻辑】如果需要，在这里处理封装类型字典
        # 备注：get_type 函数需要您在项目中实现或正确导入
        if self.package:
            self.type_dict = get_type(self.package)
        # 3. 生成缩略图
        self._generate_thumbnails()
        # 4. 根据结果更新UI界面
        if self.package:
            print(f"检测到 {len(self.package)} 个封装对象。")
            self.current = 0
            self.package_next()  # 跳转到第一个结果
        else:
            QMessageBox.information(self, '无结果', '未在该文档中发现封装图')
        # 5. 更新UI控件状态
        self._update_ui_state_after_detection(has_results=bool(self.package))

    def _generate_thumbnails(self):
        """
        辅助函数：生成检测结果的缩略图PDF。
        """
        # 从线程的属性中安全地获取有结果的页面列表
        have_page = self.thread.final_have_page if hasattr(self, 'thread') else []

        # 检查带框的PDF文件是否存在，然后再进行处理
        if not have_page or not os.path.exists(PDF_NAME):
            return

        try:
            print("正在生成缩略图...")
            ratio = 1 / 5
            with fitz.open(PDF_NAME) as doc:
                new_doc = fitz.open()
                for page_num in have_page:
                    if 0 <= page_num < doc.page_count:
                        page = doc.load_page(page_num)
                        original_rect = page.rect
                        new_rect = fitz.Rect(0, 0, original_rect.width * ratio, original_rect.height * ratio)
                        new_page = new_doc.new_page(width=new_rect.width, height=new_rect.height)
                        new_page.show_pdf_page(new_rect, doc, page_num)
                new_doc.save(PDF_NAME_MINI, garbage=4, clean=True)

            # 先加载新生成的缩略图PDF，再设置到视图中
            self.m_document_mini.load(PDF_NAME_MINI)
            self.ui.pdfView_thumbnail.setDocument(self.m_document_mini)
            print("缩略图生成完毕。")
        except Exception as e:
            print(f"生成缩略图时发生错误: {e}")
            QMessageBox.warning(self, "缩略图生成失败", f"无法生成结果缩略图。\n错误: {e}")

    def _update_ui_state_after_detection(self, has_results):
        """
        辅助函数：统一管理自动搜索结束后的UI控件状态。
        """
        self.ui.pushButton_open.setEnabled(True)
        self.ui.pushButton_detect.setEnabled(True)
        self.ui.pushButton_draw_img.setEnabled(True)
        self.ui.pushButton_draw_list.setEnabled(True)

        # 只有在有结果时，才启用“识别”和“上/下一个”按钮
        self.ui.pushButton_reco.setEnabled(has_results)
        self.ui.pushButton_pre.setEnabled(has_results)
        self.ui.pushButton_next.setEnabled(has_results)

    def img_jump(self):
        """底部封装图跳转"""
        try:
            # 获取封装图总个数
            total_img = 0
            for i in range(len(self.package)):
                # 判断数据类型
                if self.package[i]['type'] == 'img':
                    total_img += 1
            img_in = int(self.ui.lineEdit_choose_img.text())  # 用户输入跳转的符号图索引
            if (1 <= img_in <= total_img):
                # 可以跳转
                self.current = img_in
                self.judge_rect()  # 页面跳转
        except Exception as e:
            QMessageBox.critical(self, "提示", "请输入合理封装图所在索引")

    def list_jump(self):
        """底部封装表跳转"""
        try:
            # 获取封装list总个数
            total_list = 0
            total_img = 0
            for i in range(len(self.package)):
                if self.package[i]['type'] == 'list':
                    total_list += 1
                else:
                    total_img += 1
            list_in = int(self.ui.lineEdit_choose_list.text())
            if (1 <= list_in <= total_list):
                self.current = list_in + total_img
                self.judge_rect()
        except Exception as e:
            # print(e)
            QMessageBox.critical(self, "提示", "请输入合理封装表所在索引")

    def judge_rect(self):
        """
        画框
        :return:
        """
        part_content = None
        img_count = 0    # 统计封装图对象数量
        list_count = 0     # 统计手动框选表数量
        for i in range(len(self.package)):
            if self.package[i]['type'] == 'img':
                img_count += 1
            else:
                list_count += 1
        # 不同数据类型，设置画框颜色以及显示类型数量
        #   -> img 两种情况(1)单封装图  (2)封装图对象，需要部分视图
        #   -> list
        if self.package[self.current - 1]['type'] == 'img':
            self.color = (1, 0, 0)   # 框图用红色
            self.ui.lineEdit_choose_img.setText(str(self.current))
            self.ui.label_total_img.setText(IMG_LABEL.format(img_count))
            self.ui.lineEdit_choose_list.setText('—')
            self.ui.label_total_list.setText(LIST_LABEL.format(list_count))
            # self.ui.label_package.setText(LABEL.format(self.current, img_count, '-', list_count))
            part_content = self.package[self.current - 1]['part_content']
        else:
            self.color = (0, 0, 1)    # 框表用蓝色
            self.ui.lineEdit_choose_img.setText('—')
            self.ui.label_total_img.setText(IMG_LABEL.format(img_count))
            self.ui.lineEdit_choose_list.setText(str(self.current - img_count))
            self.ui.label_total_list.setText(LIST_LABEL.format(list_count))
            # self.ui.label_package.setText(LABEL.format('-', img_count, self.current - img_count, list_count))
        # self.ui.label_package.setStyleSheet(Label_initial_qss)
        # 画框
        self.draw_rect(self.package[self.current - 1]['type'], self.package[self.current - 1]['page'],
                       self.package[self.current - 1]['new_rect'], part_content)

        # 封装类型填充以及表格生成
        if self.package[self.current - 1]['package_type']:   # 有封装类型
            self.generate_table(self.package[self.current - 1]['package_type'])
            self.ui.pushButton_edit.setEnabled(1)
            self.ui.pushButton_save.setEnabled(1)
            self.ui.lineEdit_type.setText(self.package[self.current - 1]['package_type'])
        else:
            pass

        # 表格内容填充
        if self.package[self.current - 1]['reco_content'] is None:
            pass
        else:
            self.show_content()

    def draw_rect(self, rect_type, page_num, rect, part_content=None):
        """
            绘制框线
        :param rect_type:
        :param page_num:
        :param rect:
        :param part_content:
        :return:
        """
        # part_content 每一项为字典 {'part_name', 'page', 'rect', 'new_rect', }
        # keyview top side detail Form Note
        self.m_document.load(self.pdf_path)
        draw_pdf = TEMP_DIRECTORY + '\\' + r"draw.pdf"
        with fitz.open(self.pdf_path) as doc:
            page = doc[page_num]
            # 绘制框选区域
            p1 = (rect[0], rect[1])
            p2 = (rect[0], rect[3])
            p3 = (rect[2], rect[1])
            p4 = (rect[2], rect[3])
            page.draw_line(p1, p2, self.color, width=2)
            page.draw_line(p1, p3, self.color, width=2)
            page.draw_line(p3, p4, self.color, width=2)
            page.draw_line(p2, p4, self.color, width=2)
            # 框标签
            if (rect_type == 'img'):
                text = 'package'
            else:
                text = 'Form'
            page.insert_text(fitz.Point(rect[0], rect[1]), text, fontsize=12, color=self.color)
            if (part_content):     # 绘制部分视图
                for i in range(len(part_content)):
                    rect = part_content[i]['new_rect']
                    page = doc[part_content[i]['page']]
                    if (part_content[i]['part_name'] == 'Note'):
                        color = NOTE_COLOR
                        text = 'Note'
                    elif (part_content[i]['part_name'] == 'Top'):
                        color = TOP_COLOR
                        text = 'Top'
                    elif (part_content[i]['part_name'] == 'Side'):
                        color = SIDE_COLOR
                        text = 'Side'
                    elif (part_content[i]['part_name'] == 'Detail'):
                        color = DETAIL_COLOR
                        text = 'Detail'
                    elif (part_content[i]['part_name'] == 'Form'):
                        color = FORM_COLOR
                        text = 'Form'
                    else:
                        color = KEYVIEW_COLOR
                        text = part_content[i]['part_name']
                    # 绘制框选区域
                    p1 = (rect[0], rect[1])
                    p2 = (rect[0], rect[3])
                    p3 = (rect[2], rect[1])
                    p4 = (rect[2], rect[3])
                    page.draw_line(p1, p2, color=color, width=2)
                    page.draw_line(p1, p3, color=color, width=2)
                    page.draw_line(p3, p4, color=color, width=2)
                    page.draw_line(p2, p4, color=color, width=2)
                    # 框标签
                    page.insert_text(fitz.Point(rect[0], rect[1]), text, fontsize=12, color=color)
            doc.save(draw_pdf, garbage=1)
        self.temp_pdf = draw_pdf
        self.m_document.load(self.temp_pdf)
        self.nav.jump(page_num, QPoint())

        self.ui.pushButton_reco.setEnabled(1)

    def package_next(self):
        """下一个封装对象"""
        if self.current == len(self.package):
            if (self.current != 0):
                self.judge_rect()
                QMessageBox.information(self, '到达最后一项', '已经是该文档所含封装信息的最后一项')
        else:
            self.ui.pushButton_edit.setEnabled(0)  # 锁定编辑按钮
            self.ui.pushButton_save.setEnabled(0)  # 锁定保存按钮
            self.ui.tableWidget.setRowCount(0)  # 表格内容清空
            self.ui.lineEdit_type.setText('')    # 封装类型清空
            self.ui.tableWidget.setEnabled(0)
            self.ui.lineEdit_type.setEnabled(0)

            self.current += 1
            self.judge_rect()              # 获取当前封装对象信息开始画框

    def package_pre(self):
        """上一个封装对象"""
        if self.current <= 1:
            if (self.current != 0):
                self.judge_rect()
                QMessageBox.information(self, "到达最开始一项", "已经是该文档所含封装信息的最开始一项")
        else:
            self.ui.pushButton_edit.setEnabled(0)  # 锁定编辑按钮
            self.ui.pushButton_save.setEnabled(0)  # 锁定保存按钮
            self.ui.tableWidget.setRowCount(0)  # 表格内容清空
            self.ui.lineEdit_type.setText('')  # 封装类型清空
            self.ui.tableWidget.setEnabled(0)
            self.ui.lineEdit_type.setEnabled(0)

            self.current -= 1
            self.judge_rect()

    def bar_restore(self):
        # 滚动条还原
        self.ui.pdfView.horizontalScrollBar().setValue(self.page_horizontal)
        self.ui.pdfView.verticalScrollBar().setValue(self.page_vertical)

    """框图的相关设置->框图设置，框图结束设置"""
    def draw_add_func_restore(self):
        """
            框图结束，释放相关按键
        :return:
        """
        self.ui.pushButton_detect.setEnabled(1)
        self.ui.pushButton_next.setEnabled(1)
        self.ui.pushButton_pre.setEnabled(1)
        self.ui.pushButton_reco.setEnabled(1)
        self.ui.pushButton_open.setEnabled(1)
        self.ui.lineEdit_page.setEnabled(1)
        self.graphicsView.setVisible(False)            # 关闭绘图层
        self.ui.comboBox.setEnabled(1)           # 释放下拉列表
        self.ui.lineEdit_choose_list.setVisible(1)
        self.ui.lineEdit_choose_img.setVisible(1)
        self.ui.label_total_list.setVisible(1)
        self.ui.label_total_img.setVisible(1)
        self.ui.label_img_list.setVisible(0)        # 隐藏框图状态标志
        # 恢复框图，框表样式表
        self.ui.pushButton_draw_img.setStyleSheet(PuShButton_initial_qss)
        self.ui.pushButton_draw_list.setStyleSheet(PuShButton_initial_qss)

    def rect_setting_restore(self):
        """
            画完框还原布局
        :return:
        """
        if self.graphicsView.isVisible():
            self.draw_img = 0           # 重置框图标志
            self.draw_list = 0         # 重置框表标志
            self.draw_add_func_restore()     # 按钮还原
            # self.ui.label_package.setText(self.label_text)   # 恢复标签内容
            # self.ui.label_package.setStyleSheet(Label_initial_qss)

            self.bar_restore()      # 滚动条还原

    def get_img_rect_data(self):
        """
            框封装图
        :return:
        """
        if (not(self.draw_img)):
            # self.label_text = self.ui.label_package.text()
            self.draw_img = 1        # 正在框图
            self.draw_list = 0      # 取消框表
            # 开启框图标志
            self.ui.label_img_list.setVisible(1)
            self.ui.label_img_list.setText('正在框封装图...')
            self.ui.label_img_list.setStyleSheet(Label_Draw_Img_qss)
            # self.ui.label_package.setText(LABEL_DREW_IMG)   # 设置底部框图标志
            # self.ui.label_package.setStyleSheet(Label_Draw_Img_qss)  # 设置底部标签框图时样式表
            self.ui.pushButton_draw_img.setStyleSheet(PuShButton_img_Draw_qss) # 设置框图按钮样式表
            self.ui.pushButton_draw_list.setStyleSheet(PuShButton_initial_qss) # 恢复框表按钮样式表

            self.get_rect_data()    # 开始绘图
        else:     # 取消框图
            self.rect_setting_restore()     # 布局还原


    def get_list_rect_data(self):
        """
            框封装表
        :return:
        """
        if not self.draw_list:
            # self.label_text = self.ui.label_package.text()
            self.draw_list = 1      # 正在框表
            self.draw_img = 0       # 取消框图
            self.ui.label_img_list.setVisible(1)
            self.ui.label_img_list.setText('正在框封装表...')
            self.ui.label_img_list.setStyleSheet(Label_Draw_List_qss)
            # self.ui.label_package.setText(LABEL_DREW_LIST)  # 设置底部框表标志
            # self.ui.label_package.setStyleSheet(Label_Draw_List_qss)  # 设置底部标签框表时样式表
            self.ui.pushButton_draw_list.setStyleSheet(PuShButton_list_Draw_qss)  # 设置框表按钮框表时样式表
            self.ui.pushButton_draw_img.setStyleSheet(PuShButton_initial_qss)  # 恢复框图按钮样式表

            self.get_rect_data()
        else:
            self.rect_setting_restore()   # 布局还原

    def get_rect_data(self):
        """
            释放绘图层，开始框图
        :return:
        """
        # 获取水平的滚动框距离
        self.page_horizontal = self.ui.pdfView.horizontalScrollBar().value()
        # 获取竖直的滚动框距离
        self.page_vertical = self.ui.pdfView.verticalScrollBar().value()
        self.get_screen_width()        # 获取当前屏幕宽度
        # 如果屏幕宽度宽度变化了
        if self.base_screen_width != self.screen_width:
            # 清空记录
            self.pdf_factor_right_width[True].clear()
            self.pdf_factor_right_width[False].clear()
            # 重新得到pdf浏览器宽度
            self.pdf_view_width = self.ui.pdfView.size().width()
            # pdf浏览器实时高度
            self.pdf_view_height = self.ui.pdfView.size().height()
            # 保存现在的界面宽高
            self.base_screen_width = self.screen_width
        self.graphicsView.layer.setSceneRect(0, 0, self.pdf_view_width, self.pdf_view_height)  # 设置场景尺寸

        self.ui.pushButton_open.setEnabled(0)
        self.ui.pushButton_detect.setEnabled(0)
        self.ui.pushButton_pre.setEnabled(0)
        self.ui.pushButton_next.setEnabled(0)
        self.ui.pushButton_reco.setEnabled(0)
        self.ui.lineEdit_type.setEnabled(0)
        self.ui.tableWidget.setEnabled(0)
        self.ui.pushButton_edit.setEnabled(0)
        self.ui.pushButton_save.setEnabled(0)
        # 关闭跳转框
        self.ui.lineEdit_choose_list.setVisible(0)
        self.ui.lineEdit_choose_img.setVisible(0)
        self.ui.label_total_img.setVisible(0)
        self.ui.label_total_list.setVisible(0)

        self.graphicsView.setVisible(True)       # 设置画布可见
        # 跳转第二页，计算pdf显示高，只有页码大于1进行计算
        if self.pdf_page_count > 1:
            vertical_show = self.ui.pdfView.verticalScrollBar().value()
            # 刷新跳页，一定要这么跳，要不算不了真在第二页时高度
            self.nav.jump(0, QPoint())
            self.nav.jump(1, QPoint())
            # 得到pdf的展示height
            get_height = threading.Thread(target=self.get_pdf_show_height, args=(vertical_show, False,))
            get_height.start()



    def cal_rect(self, first, second, factor, is_width=True):
        """
        计算两点pdf映射
        :param first:小的点
        :param second:大的点
        :param factor:pdf展示缩放比例
        :param is_width:是否计算宽度，只有计算宽度需减去白边
        :return: 映射的两点
        """
        # 获取白边宽度
        lisp = self.pdf_factor_right_width[self.pdf_message[self.show_page_number][0]][
            factor] if is_width else 0
        # 获取两点区间，进行范围映射
        rect_range = math.ceil((second - first) * self.pdf_view_factor)
        # 计算第一个点映射
        cal_first = math.floor((first - lisp) * self.pdf_view_factor)
        # 计算第二个点映射
        cal_second = cal_first + rect_range
        # 返回结果
        return cal_first, cal_second

    @Slot(list)
    def process_rect_data(self, rect):
        """
        进行框映射，画板映射pdf
        :param rect:需要映射的框
        :return:
        """
        # 框图结束，释放相关按键以及隐藏绘图层
        self.draw_add_func_restore()
        # 获取pdf页码
        self.show_page_number = self.nav.currentPage()
        # 水平方向距离
        self.page_horizontal = self.ui.pdfView.horizontalScrollBar().value()
        # 获取竖直的滚动框距离
        self.page_vertical = self.ui.pdfView.verticalScrollBar().value()
        # 处理添加浏览框水平垂直距离
        rect[0] += self.page_horizontal
        rect[2] += self.page_horizontal
        # 只有页码大于1进行计算
        if self.pdf_page_count > 1:
            # 计算下拉整页距离，竖向页面加横向页面{页码:[是否颠倒，宽，高，宽高比，前面几个颠倒页]}
            cal_vertical = (self.show_page_number - self.pdf_message[self.show_page_number][4]) * self.pdf_show_height + \
                           self.pdf_message[self.show_page_number][4] * self.trans_pdf_show_width + \
                           self.ui.pdfView.pageSpacing() * self.show_page_number + self.pdf_margin
        else:
            # 只有一页
            cal_vertical = self.pdf_margin
        rect[1] += self.page_vertical - cal_vertical
        rect[3] += self.page_vertical - cal_vertical
        # 如果没有存储白边
        if self.pdf_factor_right_width[self.pdf_message[self.show_page_number][0]].get(self.pdf_factor) is None:
            # 计算当前pdf展示宽高
            self.document_screen_width = self.pdf_view_width - self.ui.pdfView.verticalScrollBar().width()
            # 如果没有颠倒页或者就是颠倒页,则pdf的放大缩小两边边框相同，不用截图直接计算
            if self.pdf_have_trans is False or self.pdf_message[self.show_page_number][0]:
                # 计算使用的宽度
                now_pdf_width = self.pdf_show_width if self.pdf_message[self.show_page_number][
                                                           0] is False else self.trans_pdf_show_height
                # 如果显示大小超过pdf浏览器屏幕的宽
                if now_pdf_width > self.document_screen_width:
                    lisp = self.pdf_margin
                else:
                    # 没有超过浏览器的宽度，则平均计算
                    lisp = (self.document_screen_width - now_pdf_width) / 2
            # 有横页并且不是横页
            else:
                # 根据pdf的高度判断是否影响宽度的计算,如果显示的高度大于屏幕的宽度
                if self.trans_pdf_show_height > self.document_screen_width:
                    lisp = (self.trans_pdf_show_height - self.pdf_show_width) / 2 + self.pdf_margin
                else:
                    # 没有超过浏览器的宽度，则取宽度平均计算
                    lisp = (self.document_screen_width - self.pdf_show_width) / 2
            # 存储结果
            self.pdf_factor_right_width[self.pdf_message[self.show_page_number][0]][
                self.pdf_factor] = lisp
        # 处理x方向数据
        rect[0], rect[2] = self.cal_rect(rect[0], rect[2], self.pdf_factor)
        # 处理y方向数据
        rect[1], rect[3] = self.cal_rect(rect[1], rect[3], self.pdf_factor, False)

        # 界限数据,需要判断是否颠倒
        rect[0] = max(rect[0], 0)
        rect[2] = min(rect[2], self.pdf_message[self.show_page_number][1])
        rect[1] = max(rect[1], 0)
        rect[3] = min(rect[3], self.pdf_message[self.show_page_number][2])

        # 处理数据(图片坐标系到pdf坐标系) 框图还是框表 -> 生成数据 -> 排序 -> 绘制框线
        # 若是封装图'img' ->{'page', 'type', 'rect', 'new_rect', 'package_type', 'part_content': [{}, {}], 'reco_content'}
        #         -> {'page', 'type', 'rect', 'new_rect', 'package_type', 'part_content': [], 'reco_content'}
        # 若是封装表'list' ->{'page', 'type', 'rect', 'new_rect', 'reco_content'}
        # 处理数据
        new_rect = [0, 0, 0, 0]
        with fitz.open(self.pdf_path) as doc:
            rotation = doc[self.show_page_number].rotation     # 页面旋转角度
            page_width = doc[self.show_page_number].rect[2]
            page_height = doc[self.show_page_number].rect[3]
        if (rotation == 90):
            new_rect[0] = rect[1]
            new_rect[1] = page_width - rect[0]
            new_rect[2] = rect[3]
            new_rect[3] = page_width - rect[2]
        elif (rotation == 0):
            new_rect[0] = rect[0]
            new_rect[1] = rect[1]
            new_rect[2] = rect[2]
            new_rect[3] = rect[3]
        else:
            new_rect = rect
            QMessageBox.information(self, '提示信息', "天阶夜色凉如水\n卧看牵牛织女星")
        # 框图还是框表 生成数据
        if (self.draw_img):       # 框图数据
            data_dict = {'page': self.show_page_number, 'type': 'img',
                        'rect': rect, 'new_rect': new_rect, 'package_type': None, 'source': 'manual',
                        'part_content': None, 'reco_content': None}
        else:             # 框表数据
            data_dict = {'page': self.show_page_number, 'type': 'list',
                        'rect': rect, 'new_rect': new_rect, 'package_type': None, 'source': 'manual',
                        'part_content': None, 'reco_content': None}
        self.package.append(data_dict)

        self.draw_img = 0   # 恢复框图标志
        self.draw_list = 0   # 恢复框表标志

        # 对package信息进行排序
        self.package = sorted(self.package, key=lambda i: (i['type'], i['page']))
        self.current = self.package.index(data_dict)
        self.package_next()

        self.bar_restore()


    """pdf浏览器的附属功能->页面展示，页面跳转"""
    def show_page(self):
        """
        更新pdf浏览器页码标签
        """
        self.show_page_number = self.nav.currentPage()
        self.ui.lineEdit_page.setText(str(self.show_page_number + 1))  # pdf浏览器页数改变时标签栏也改变

    def pdf_combo_show(self):
        """
        pdf浏览器下拉列表功能选择
        """
        # 获取当前索引值
        index = self.ui.comboBox.currentIndex()
        if index == 0:  # 下拉列表显示百分比
            # 设置百分比
            self.set_factor()
            # 开放放大缩小按钮
            self.ui.pushButton_zoomin.setEnabled(1)
            self.ui.pushButton_zoomout.setEnabled(1)

    def set_factor(self):
        """
        设置缩放比
        :return:
        """
        # 计算保持原来的页码
        save_v = self.ui.pdfView.verticalScrollBar().value()
        # 获取当前页码
        self.show_page_number = self.nav.currentPage()
        # 获取下来框数目
        count = self.ui.comboBox.count()
        # 设置百分比
        self.ui.pdfView.setZoomFactor(self.pdf_factor)
        # 计算滚动条
        vertical_show = (save_v - self.pdf_margin - self.show_page_number * self.ui.pdfView.pageSpacing()) * \
                        self.in_or_out + self.pdf_margin + self.show_page_number * self.ui.pdfView.pageSpacing()
        # 跳转到计算的滚动位置
        self.ui.pdfView.verticalScrollBar().setValue(vertical_show)
        # 显示百分比
        if count > len(COMO_ITEMS):
            self.ui.comboBox.setItemText(count - 1, str(int(self.pdf_factor * 100)) + '%')
        else:
            self.ui.comboBox.addItem(str(int(self.pdf_factor * 100)) + '%')
        self.ui.comboBox.setCurrentIndex(len(COMO_ITEMS))

    def pdf_zoom_out(self):
        """
        pdf浏览器界面放大
        :return:
        """
        # 获取放大比例
        self.pdf_factor = self.pdf_factor * ZOOM_MULTIPLIER  # 放大比例
        # 获取滚动计算比例
        self.in_or_out = ZOOM_MULTIPLIER
        self.set_factor()

    def pdf_zoom_in(self):
        """
        pdf浏览器界面缩小
        :return:
        """
        # 获取缩小比例
        self.pdf_factor = self.pdf_factor / ZOOM_MULTIPLIER  # 缩小比例
        # 获取滚动计算比例
        self.in_or_out = 1 / ZOOM_MULTIPLIER
        self.set_factor()

    def page_jump_restore(self, page):
        """
        指定页码进行跳转
        """
        # 保存跳转页码
        self.show_page_number = page
        # 进行跳转
        self.nav.jump(page, QPoint())  # pdf页面跳转

    def edit_page_jump(self):
        """
        获取输入框页码跳转
        :return:
        """
        # 获取页码 实际页码
        page = int(self.ui.lineEdit_page.text())
        # 判断范围
        if 0 < page <= self.m_document.pageCount():
            self.page_jump_restore(page - 1)
        else:  # 超出范围,不跳
            self.ui.lineEdit_page.setText(str(self.show_page_number + 1))

    """窗口动态事件"""
    def closeEvent(self, event: QCloseEvent):
        """关闭系统"""
        enquire = EnquirePopUp(self, '退出程序', '确认要退出吗？')
        if (enquire.enquire_result == Yes):
            try:
                self.m_document.load(self.ace_pdf)
                self.m_document_mini.load(self.ace_pdf)
            except Exception as e:
                pass
            try:
                remove_dir(TEMP_DIRECTORY)
            except Exception as e:
                pass
            event.accept()
        else:
            event.ignore()

    def resizeEvent(self, event):
        """窗口大小变化时，更新场景大小，更新pdf浏览器大小和显示表格大小"""
        self.pdf_view_width = self.ui.pdfView.size().width()     # pdf浏览器实时宽度
        self.pdf_view_height = self.ui.pdfView.size().height()   # pdf浏览器实时高度

        self.pdf_view_width_mini = self.ui.pdfView_thumbnail.size().width()

        # 规范参数表格表头宽度
        self.table_width = self.ui.tableWidget.width()    # 表格实时宽度
        width = (self.table_width - TABLE_GAP) / len(TABLE_HEADERS)
        for i in range(len(TABLE_HEADERS)):
            self.ui.tableWidget.setColumnWidth(i, width)

        # 更新场景大小
        if self.graphicsView.isVisible():
            self.graphicsView.layer.setSceneRect(0, 0, self.pdf_view_width, self.pdf_view_height)



if __name__ == '__main__':
    app = QApplication(sys.argv)
    myWindow = MyWindow()
    myWindow.showMaximized()
    sys.exit(app.exec())